"""A normalised description of a workbook, so two writers can be compared.

The browser version writes its xlsx by hand rather than through openpyxl, so
the two files can never be byte-identical: they order zip entries differently,
express the same style in a different shape, and disagree about whitespace no
reader ever sees. Comparing bytes would fail on every run while proving
nothing, so what gets compared instead is what a person opening the file would
find: which sheets exist, what is in each cell, how it is typed and formatted,
what is merged, how wide the columns are and where the freeze sits.

Fonts, fills, colours and alignment are deliberately left out. That is a scope
decision, not an oversight. The two implementations model styling completely
differently -- openpyxl hangs a style object off each cell, the browser writer
registers a look once and stores an index -- so comparing them would compare
two encodings rather than two workbooks. The facts below are the ones that
decide what a cell actually shows.
"""

from __future__ import annotations

from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple


def _freeze_row(sheet: Any) -> int:
    """The last frozen row, or 0 where nothing is frozen.

    openpyxl states this as the top-left cell of the *unfrozen* region -- "A4"
    means rows 1 to 3 are frozen -- which is one off from the count anything
    comparing two sheets wants.
    """
    frozen = sheet.freeze_panes
    if not frozen:
        return 0
    return coordinate_to_tuple(str(frozen))[0] - 1


def _column_widths(sheet: Any) -> list[float]:
    """One width per column the sheet uses, in column order."""
    return [
        round(sheet.column_dimensions[get_column_letter(index)].width, 2)
        for index in range(1, sheet.max_column + 1)
    ]


def _cell_facts(cell: Any) -> dict[str, Any]:
    value = cell.value
    if isinstance(value, str):
        return {
            "row": cell.row,
            "col": cell.column,
            "value": value,
            "type": "s",
            "numberFormat": cell.number_format,
        }

    # Six decimals is far beyond anything the table carries and short of where
    # float64 starts showing its own representation noise, which is the only
    # difference two languages computing the same number produce. The zero
    # case flattens -0.0, which reads as 0 but is not equal to it under the
    # strict deep-equal on the other side of the comparison.
    number = round(float(value), 6)
    return {
        "row": cell.row,
        "col": cell.column,
        "value": 0.0 if number == 0 else number,
        "type": "n",
        "numberFormat": cell.number_format,
    }


def _sheet_facts(sheet: Any) -> dict[str, Any]:
    cells = [
        _cell_facts(cell)
        for row in sheet.iter_rows()
        for cell in row
        # An empty cell is written only to carry a fill, and fills are not
        # compared, so it says nothing about what the sheet contains.
        if cell.value is not None and cell.value != ""
    ]
    cells.sort(key=lambda cell: (cell["row"], cell["col"]))

    return {
        "name": sheet.title,
        "freezeRow": _freeze_row(sheet),
        "columnWidths": _column_widths(sheet),
        "merges": sorted(str(area) for area in sheet.merged_cells.ranges),
        "cells": cells,
    }


def workbook_facts(source: Any) -> dict[str, Any]:
    """Normalised description of a workbook, for comparing two writers.

    ``source`` is a path, an open binary stream, or an openpyxl Workbook.
    """
    book = source if isinstance(source, Workbook) else load_workbook(source)
    return {"sheets": [_sheet_facts(book[name]) for name in book.sheetnames]}
