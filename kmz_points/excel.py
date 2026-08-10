"""Writing the table out as a workbook."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import BinaryIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from kmz_points.table import (
    COLUMNS,
    build_table_rows,
    NUMBER_INDEX,
    SOURCE_FILE_INDEX,
    bands,
)

# Hard limit imposed by the xlsx format; a longer string makes the file
# unopenable rather than merely ugly.
EXCEL_CELL_LIMIT = 32767

_MIN_WIDTH = 8
_MAX_WIDTH = 60
_WIDTH_PADDING = 2

# Three header rows: band titles, band captions, then the column names. Data
# starts at row 4, and a grey banner naming the source file precedes each
# file's rows.
BAND_TITLE_ROW = 1
BAND_CAPTION_ROW = 2
HEADER_ROW = 3
FIRST_DATA_ROW = 4

_BANNER_FILL = "A6A6A6"
# Holes sit inside an area, so their banner is lighter than the area's own --
# the indentation of colour rather than of whitespace, which a merged row
# cannot show.
_HOLE_BANNER_FILL = "D9D9D9"
_CENTRED = Alignment(horizontal="center", vertical="center")


def _fill(colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=colour)


def data_rows(sheet) -> list[tuple]:
    """The point rows of a Points sheet, skipping headers and file banners.

    max_row is no longer the point count: the sheet carries three header rows
    and one banner per source file. A banner sets only its first cell, so the
    numbered "#" column tells the two apart.
    """
    return [
        row
        for row in sheet.iter_rows(min_row=FIRST_DATA_ROW)
        if row[NUMBER_INDEX].value is not None
    ]


def output_filename(when: datetime | None = None) -> str:
    """``points_YYYYMMDD_HHMMSS.xlsx`` for the given moment (default: now).

    Seconds are included because minutes are not enough: two exports in the
    same minute produced the same name, and on the desktop the second silently
    overwrote the first while both reported success.
    """
    when = when or datetime.now()
    return f"points_{when:%Y%m%d_%H%M%S}.xlsx"


def unique_path(directory: str | Path, filename: str) -> Path:
    """A path in ``directory`` that does not already exist.

    Seconds make a collision unlikely, not impossible -- two exports can still
    land in the same second, and a clock can go backwards. Overwriting someone's
    workbook is bad enough to be worth ruling out rather than making improbable,
    so a name already taken gains a counter: ``points_….xlsx``,
    ``points_…-2.xlsx``, and so on.
    """
    directory = Path(directory)
    candidate = directory / filename
    if not candidate.exists():
        return candidate

    stem = Path(filename).stem
    suffix = Path(filename).suffix
    attempt = 2
    while True:
        candidate = directory / f"{stem}-{attempt}{suffix}"
        if not candidate.exists():
            return candidate
        attempt += 1


# Only "=" needs defusing in an xlsx. openpyxl infers a type from the string
# it is handed, and a leading "=" becomes data_type "f" -- a live formula
# built from content we did not write. point.name and point.description come
# straight from KML, and an uploaded filename reaches the Issues sheet, so
# the guard lives here rather than at any single caller.
#
# "+", "-" and "@" are deliberately NOT included. They matter when a user
# types them into a cell, or when a CSV is imported, but a string written
# into an xlsx stays a string: openpyxl gives all three data_type "s". Adding
# them cost real data instead -- a placemark called "-Alpha" or a description
# holding "+44 7700 900000" was rewritten with a leading apostrophe that then
# travelled with every copy, sort and re-import.
_FORMULA_TRIGGER = "="


def _fit(value):
    """Clamp a cell value to what Excel will accept, and stop a leading "="
    from being turned into a live formula."""
    if not isinstance(value, str):
        return value
    if value.startswith(_FORMULA_TRIGGER):
        # A leading apostrophe makes openpyxl store the rest as literal text.
        value = "'" + value
    if len(value) > EXCEL_CELL_LIMIT:
        return value[: EXCEL_CELL_LIMIT - 3] + "..."
    return value


def _column_width(header: str, values: list) -> float:
    longest = max(
        [len(header)] + [len(str(v)) for v in values if v is not None],
        default=len(header),
    )
    return min(max(longest + _WIDTH_PADDING, _MIN_WIDTH), _MAX_WIDTH)


def _write_header(sheet) -> None:
    """Band titles, band captions, then the column names."""
    for band, start, end in bands():
        first = get_column_letter(start + 1)
        last = get_column_letter(end + 1)

        title = sheet.cell(row=BAND_TITLE_ROW, column=start + 1, value=band.title)
        title.font = Font(bold=True, size=12)
        title.alignment = _CENTRED
        if band.fill:
            title.fill = _fill(band.fill)

        caption = sheet.cell(row=BAND_CAPTION_ROW, column=start + 1)
        if band.caption:
            caption.value = band.caption
            caption.alignment = _CENTRED
            if band.caption_fill:
                caption.fill = _fill(band.caption_fill)
            sheet.merge_cells(f"{first}{BAND_CAPTION_ROW}:{last}{BAND_CAPTION_ROW}")
            sheet.merge_cells(f"{first}{BAND_TITLE_ROW}:{last}{BAND_TITLE_ROW}")
        else:
            # No caption, so the title claims both rows rather than leaving a
            # blank band beneath it.
            sheet.merge_cells(
                start_row=BAND_TITLE_ROW,
                start_column=start + 1,
                end_row=BAND_CAPTION_ROW,
                end_column=end + 1,
            )
            if band.fill:
                caption.fill = _fill(band.fill)

    for index, column in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=HEADER_ROW, column=index, value=column.header)
        cell.font = Font(bold=True)
        cell.alignment = _CENTRED
        if column.fill:
            cell.fill = _fill(column.fill)
        if column.font_colour:
            cell.font = Font(bold=True, color=column.font_colour)


def _banner(sheet, row_number: int, text: str, colour: str) -> None:
    """A merged, filled row spanning the whole table."""
    cell = sheet.cell(row=row_number, column=1, value=_fit(text))
    cell.font = Font(bold=True, color="FFFFFF" if colour == _BANNER_FILL else "1F2933")
    cell.alignment = _CENTRED
    cell.fill = _fill(colour)
    sheet.merge_cells(
        start_row=row_number,
        start_column=1,
        end_row=row_number,
        end_column=len(COLUMNS),
    )


def _area_banner_text(measured) -> str:
    """What an area's banner says: its size, or why there isn't one."""
    area = measured.area
    name = area.name or "<unnamed>"
    corners = f"{area.corner_count} corners"
    size = measured.measurement

    if size.square_metres is None:
        return f"{name} — area not measured: {size.problem} · {corners}"

    return (
        f"{name} — {size.square_metres:,.0f} m² · "
        f"{size.hectares:,.3f} ha · {size.square_kilometres:,.6f} km² · {corners}"
    )


def _write_body(sheet, rows: list[list]) -> list[int]:
    """Write the rows, banner-separated by source file.

    Returns the row numbers holding data, so number formats skip the banners.
    Grouping walks consecutive runs rather than sorting: points already arrive
    in file order, and sorting would renumber the batch out of sequence.
    """
    data_rows: list[int] = []
    current_file = None
    row_number = FIRST_DATA_ROW

    for row in rows:
        source = row[SOURCE_FILE_INDEX]
        if source != current_file:
            current_file = source
            banner = sheet.cell(row=row_number, column=1, value=_fit(source))
            banner.font = Font(bold=True, color="FFFFFF")
            banner.alignment = _CENTRED
            banner.fill = _fill(_BANNER_FILL)
            sheet.merge_cells(
                start_row=row_number,
                start_column=1,
                end_row=row_number,
                end_column=len(COLUMNS),
            )
            row_number += 1

        for index, value in enumerate(row, start=1):
            sheet.cell(row=row_number, column=index, value=_fit(value))
        data_rows.append(row_number)
        row_number += 1

    return data_rows


def _apply_formats(sheet, data_row_numbers: list[int], rows: list[list]) -> None:
    """Number formats on data rows only, and a width per column."""
    for index, column in enumerate(COLUMNS, start=1):
        letter = get_column_letter(index)
        if column.number_format:
            for row_number in data_row_numbers:
                sheet.cell(row=row_number, column=index).number_format = (
                    column.number_format
                )
        sheet.column_dimensions[letter].width = _column_width(
            column.header, [row[index - 1] for row in rows]
        )


def _write_areas(book, measured_areas: list) -> None:
    """The Areas sheet: a banner per area, its corners, then each hole.

    Corner rows are ordinary point rows, so they carry every conversion the
    Points sheet does. Numbering restarts within each ring, which is what makes
    a corner list readable.
    """
    sheet = book.create_sheet("Areas")
    _write_header(sheet)

    data_row_numbers: list[int] = []
    all_rows: list[list] = []
    row_number = FIRST_DATA_ROW

    for measured in measured_areas:
        _banner(sheet, row_number, _area_banner_text(measured), _BANNER_FILL)
        row_number += 1

        rings = [(None, measured.area.outer)]
        for position, hole in enumerate(measured.area.holes, start=1):
            rings.append((f"hole {position} — {len(hole)} corners", hole))

        for label, corners in rings:
            if label is not None:
                _banner(sheet, row_number, label, _HOLE_BANNER_FILL)
                row_number += 1

            for row in build_table_rows(corners):
                all_rows.append(row)
                for index, value in enumerate(row, start=1):
                    sheet.cell(row=row_number, column=index, value=_fit(value))
                data_row_numbers.append(row_number)
                row_number += 1

    sheet.freeze_panes = f"A{FIRST_DATA_ROW}"
    _apply_formats(sheet, data_row_numbers, all_rows)


def write_workbook(
    rows: list[list],
    target: str | Path | BinaryIO,
    issues: list[str] | None = None,
    areas: list | None = None,
) -> Path | None:
    """Write rows to an xlsx file, or into an open binary stream.

    The web service needs the workbook in memory, so a stream is accepted as
    well as a path. Returns the path written, or None for a stream.

    ``issues`` adds a second sheet naming files that could not be read. A
    browser download has nowhere else to report them: the response body is the
    workbook. The desktop app passes nothing and its output is unchanged.
    """
    book = Workbook()
    sheet = book.active
    sheet.title = "Points"

    _write_header(sheet)
    written = _write_body(sheet, rows)
    sheet.freeze_panes = f"A{FIRST_DATA_ROW}"
    _apply_formats(sheet, written, rows)

    if areas:
        # Between Points and Issues: the data first, the complaints last.
        _write_areas(book, areas)

    if issues:
        # Appended after Points, so opening the file lands on the data.
        notes = book.create_sheet("Issues")
        notes.append(["Issue"])
        notes["A1"].font = Font(bold=True)
        for line in issues:
            notes.append([_fit(line)])
        notes.column_dimensions["A"].width = _MAX_WIDTH

    if isinstance(target, (str, Path)):
        path = Path(target)
        path.parent.mkdir(parents=True, exist_ok=True)
        book.save(path)
        return path

    book.save(target)
    return None
