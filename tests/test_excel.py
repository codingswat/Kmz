"""Excel writer tests -- every assertion is made against a reopened workbook."""

import io
from datetime import datetime

import openpyxl
import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from kmz_points.excel import (
    BAND_CAPTION_ROW,
    BAND_TITLE_ROW,
    EXCEL_CELL_LIMIT,
    FIRST_DATA_ROW,
    HEADER_ROW,
    data_rows,
    output_filename,
    unique_path,
    write_workbook,
)
from kmz_points.models import Point
from kmz_points.table import (
    COMBINED,
    SEPARATED,
    SEPARATION,
    bands,
    build_table_rows,
    column_index,
    headers,
)


def make_point(
    alt=120.5,
    description="desc",
    lat=34.567890,
    lon=38.123456,
    name="Alpha",
    source="a.kml",
):
    return Point(
        name=name,
        description=description,
        lon=lon,
        lat=lat,
        alt=alt,
        source_file=source,
    )


def write_and_reopen(tmp_path, points):
    path = tmp_path / "out.xlsx"
    write_workbook(build_table_rows(points), path)
    return openpyxl.load_workbook(path)


def first_data_row(sheet):
    """The cells of the first point row, with headers and file banners
    already skipped by kmz_points.excel.data_rows."""
    return data_rows(sheet)[0]


class TestOutputFilename:
    def test_uses_the_specified_pattern(self):
        stamp = datetime(2026, 8, 6, 14, 35)
        assert output_filename(stamp) == "points_20260806_143500.xlsx"

    def test_pads_single_digit_components(self):
        stamp = datetime(2026, 1, 2, 3, 4)
        assert output_filename(stamp) == "points_20260102_030400.xlsx"


class TestHeaderRows:
    """Three header rows now: band titles, band captions, then column
    names. Data starts at FIRST_DATA_ROW."""

    def test_column_names_are_on_the_header_row(self, tmp_path):
        sheet = write_and_reopen(tmp_path, [make_point()]).active
        written = [c.value for c in sheet[HEADER_ROW]]
        assert written == headers()

    def test_header_row_is_bold(self, tmp_path):
        sheet = write_and_reopen(tmp_path, [make_point()]).active
        assert all(c.font.bold for c in sheet[HEADER_ROW])

    def test_sheet_is_frozen_below_the_header_rows(self, tmp_path):
        sheet = write_and_reopen(tmp_path, [make_point()]).active
        assert sheet.freeze_panes == f"A{FIRST_DATA_ROW}"

    def test_column_widths_are_set(self, tmp_path):
        sheet = write_and_reopen(tmp_path, [make_point()]).active
        widths = [d.width for d in sheet.column_dimensions.values()]
        assert len(widths) == len(headers())
        assert all(w and w > 0 for w in widths)

    def test_band_titles_appear_on_the_first_header_row(self, tmp_path):
        sheet = write_and_reopen(tmp_path, [make_point()]).active
        for band, start, _end in bands():
            assert sheet.cell(row=BAND_TITLE_ROW, column=start + 1).value == band.title

    def test_only_separation_carries_a_caption(self, tmp_path):
        # "separation" is the only band with a second-row caption ("decimal
        # degrees"); every other band's title claims both header rows
        # instead, so its caption-row cell has nothing of its own.
        sheet = write_and_reopen(tmp_path, [make_point()]).active
        for band, start, _end in bands():
            value = sheet.cell(row=BAND_CAPTION_ROW, column=start + 1).value
            if band is SEPARATION:
                assert value == "decimal degrees"
            else:
                assert value is None

    def test_separation_title_and_caption_merge_as_two_separate_bands(self, tmp_path):
        sheet = write_and_reopen(tmp_path, [make_point()]).active
        _band, start, end = next(b for b in bands() if b[0] is SEPARATION)
        first, last = get_column_letter(start + 1), get_column_letter(end + 1)
        ranges = {str(r) for r in sheet.merged_cells.ranges}
        assert f"{first}{BAND_TITLE_ROW}:{last}{BAND_TITLE_ROW}" in ranges
        assert f"{first}{BAND_CAPTION_ROW}:{last}{BAND_CAPTION_ROW}" in ranges

    def test_a_captionless_band_merges_its_title_down_across_both_rows(self, tmp_path):
        sheet = write_and_reopen(tmp_path, [make_point()]).active
        _band, start, end = next(b for b in bands() if b[0] is COMBINED)
        first, last = get_column_letter(start + 1), get_column_letter(end + 1)
        expected = f"{first}{BAND_TITLE_ROW}:{last}{BAND_CAPTION_ROW}"
        assert expected in {str(r) for r in sheet.merged_cells.ranges}


class TestCellTypes:
    def test_decimal_degrees_are_stored_as_numbers(self, tmp_path):
        sheet = write_and_reopen(tmp_path, [make_point()]).active
        cell = first_data_row(sheet)[column_index("latitude", band=SEPARATION)]
        assert isinstance(cell.value, float)
        assert cell.value == pytest.approx(34.567890)

    def test_decimal_degrees_display_six_places(self, tmp_path):
        sheet = write_and_reopen(tmp_path, [make_point()]).active
        cell = first_data_row(sheet)[column_index("latitude", band=SEPARATION)]
        assert cell.number_format == "0.000000"

    def test_easting_is_stored_as_a_number(self, tmp_path):
        sheet = write_and_reopen(tmp_path, [make_point()]).active
        cell = first_data_row(sheet)[column_index("Easting (m)")]
        assert isinstance(cell.value, int)
        assert cell.value == 419595

    def test_elevation_is_stored_as_a_number(self, tmp_path):
        sheet = write_and_reopen(tmp_path, [make_point()]).active
        cell = first_data_row(sheet)[column_index("elevation", band=SEPARATION)]
        assert cell.value == pytest.approx(120.5)

    def test_combined_dms_column_is_stored_as_text(self, tmp_path):
        sheet = write_and_reopen(tmp_path, [make_point()]).active
        cell = first_data_row(sheet)[column_index("latitude", band=COMBINED)]
        assert isinstance(cell.value, str)
        assert cell.value == "34° 34' 4.40\" N"

    def test_missing_altitude_leaves_the_cell_empty(self, tmp_path):
        sheet = write_and_reopen(tmp_path, [make_point(alt=None)]).active
        cell = first_data_row(sheet)[column_index("elevation", band=SEPARATION)]
        assert cell.value is None

    def test_separated_degrees_are_stored_as_magnitudes(self, tmp_path):
        # Companion to the table-level test: confirms the sign-stripping
        # survives the trip through openpyxl and back, not just
        # build_table_rows. Uses the same latitude as the "Bravo" sample
        # point: south, but with a whole-degree part of 0, which has no sign
        # of its own -- the hemisphere has to live in the decimal instead.
        sheet = write_and_reopen(
            tmp_path, [make_point(lat=-0.180653, lon=-78.467834)]
        ).active
        row = first_data_row(sheet)
        lat = column_index("lat", band=SEPARATED)

        assert row[lat].value == pytest.approx(-0.180653)
        assert (row[lat + 1].value, row[lat + 2].value, row[lat + 3].value) == (
            0,
            10,
            50.35,
        )


class TestOversizedContent:
    def test_description_longer_than_excel_allows_is_truncated(self, tmp_path):
        sheet = write_and_reopen(tmp_path, [make_point(description="x" * 40000)]).active
        cell = first_data_row(sheet)[column_index("Description")]
        assert len(cell.value) <= EXCEL_CELL_LIMIT

    def test_truncation_is_marked_rather_than_silent(self, tmp_path):
        sheet = write_and_reopen(tmp_path, [make_point(description="x" * 40000)]).active
        cell = first_data_row(sheet)[column_index("Description")]
        assert cell.value.endswith("...")


class TestFormulaInjection:
    """openpyxl infers a type from the string it is handed, and a leading "="
    becomes a live formula built from content we did not write. point.name and
    point.description come straight from KML, so an attacker controlling a
    placemark's name controls what lands in the workbook."""

    def test_a_name_starting_with_equals_lands_as_text_not_a_formula(self, tmp_path):
        point = make_point(name='=HYPERLINK("http://evil.example","Click")+A1')
        sheet = write_and_reopen(tmp_path, [point]).active
        cell = first_data_row(sheet)[column_index("Name")]
        assert cell.data_type != "f"
        assert cell.value.startswith("'=")

    def test_an_ordinary_name_is_unchanged(self, tmp_path):
        sheet = write_and_reopen(tmp_path, [make_point(name="Alpha")]).active
        cell = first_data_row(sheet)[column_index("Name")]
        assert cell.value == "Alpha"

    @pytest.mark.parametrize("name", ["-Alpha", "+44 7700 900000", "@site", "@1"])
    def test_other_leading_symbols_are_left_alone(self, tmp_path, name):
        # These matter when a user types them into a cell or imports a CSV,
        # but a string written into an xlsx stays a string. Escaping them
        # corrupted ordinary content -- bulleted names, phone numbers -- with
        # an apostrophe that then travelled with every copy and re-import.
        sheet = write_and_reopen(tmp_path, [make_point(name=name)]).active
        cell = first_data_row(sheet)[column_index("Name")]
        assert cell.value == name
        assert cell.data_type != "f"

    def test_a_leading_minus_is_still_a_string_not_a_formula(self, tmp_path):
        # The claim the previous escaping rested on, checked rather than
        # assumed: openpyxl does not turn these into formulas.
        sheet = write_and_reopen(tmp_path, [make_point(name="-2+3")]).active
        assert first_data_row(sheet)[column_index("Name")].data_type == "s"


class TestRows:
    def test_every_point_gets_a_row(self, tmp_path):
        book = write_and_reopen(tmp_path, [make_point(), make_point(), make_point()])
        assert len(data_rows(book.active)) == 3

    def test_empty_input_still_writes_a_header(self, tmp_path):
        book = write_and_reopen(tmp_path, [])
        assert [c.value for c in book.active[HEADER_ROW]] == headers()
        assert data_rows(book.active) == []
        assert book.active.max_row == HEADER_ROW


class TestSourceFileBanners:
    """A grey merged banner names the source file immediately above each
    group of its points."""

    def test_a_banner_precedes_each_files_group(self, tmp_path):
        points = [
            make_point(name="Alpha", source="a.kml"),
            make_point(name="Bravo", source="a.kml"),
            make_point(name="Charlie", source="b.kml"),
        ]
        sheet = write_and_reopen(tmp_path, points).active
        rows = data_rows(sheet)
        assert len(rows) == 3

        first_of_a = rows[0][0].row
        first_of_b = rows[2][0].row
        assert sheet.cell(row=first_of_a - 1, column=1).value == "a.kml"
        assert sheet.cell(row=first_of_b - 1, column=1).value == "b.kml"

    def test_a_shared_source_file_gets_one_banner_not_one_per_point(self, tmp_path):
        points = [make_point(source="a.kml"), make_point(source="a.kml")]
        sheet = write_and_reopen(tmp_path, points).active
        rows = data_rows(sheet)
        # Nothing (i.e. no second banner) sits between the two data rows.
        assert rows[1][0].row == rows[0][0].row + 1

    def test_banner_rows_are_never_counted_as_data(self, tmp_path):
        # Three files, three points, three banners -- data_rows must still
        # report exactly three points, not six.
        points = [
            make_point(source="a.kml"),
            make_point(source="b.kml"),
            make_point(source="c.kml"),
        ]
        sheet = write_and_reopen(tmp_path, points).active
        assert len(data_rows(sheet)) == 3

    def test_the_banner_spans_every_column(self, tmp_path):
        sheet = write_and_reopen(tmp_path, [make_point(source="a.kml")]).active
        banner_row = data_rows(sheet)[0][0].row - 1
        first, last = get_column_letter(1), get_column_letter(len(headers()))
        expected = f"{first}{banner_row}:{last}{banner_row}"
        assert expected in {str(r) for r in sheet.merged_cells.ranges}


class TestWritingToAStream:
    """The web service needs a workbook in memory, never on disk."""

    def test_a_stream_target_produces_a_readable_workbook(self):
        buffer = io.BytesIO()
        write_workbook(build_table_rows([make_point()]), buffer)
        buffer.seek(0)
        sheet = load_workbook(buffer).active
        assert [c.value for c in sheet[HEADER_ROW]] == headers()
        assert len(data_rows(sheet)) == 1

    def test_a_stream_target_returns_no_path(self):
        assert write_workbook(build_table_rows([make_point()]), io.BytesIO()) is None

    def test_a_path_target_still_returns_its_path(self, tmp_path):
        target = tmp_path / "out.xlsx"
        assert write_workbook(build_table_rows([make_point()]), target) == target


class TestIssuesSheet:
    """A browser download has nowhere to show a warning, so failures ride
    along inside the workbook."""

    def test_no_issues_means_no_second_sheet(self):
        buffer = io.BytesIO()
        write_workbook(build_table_rows([make_point()]), buffer)
        buffer.seek(0)
        assert load_workbook(buffer).sheetnames == ["Points"]

    def test_issues_are_listed_on_their_own_sheet(self):
        buffer = io.BytesIO()
        write_workbook(
            build_table_rows([make_point()]),
            buffer,
            issues=["broken.kmz: not a readable KMZ archive"],
        )
        buffer.seek(0)
        book = load_workbook(buffer)
        assert book.sheetnames == ["Points", "Issues"]
        listed = [row[0].value for row in book["Issues"].iter_rows(min_row=2)]
        assert listed == ["broken.kmz: not a readable KMZ archive"]

    def test_the_points_sheet_stays_first(self):
        buffer = io.BytesIO()
        write_workbook(build_table_rows([make_point()]), buffer, issues=["a problem"])
        buffer.seek(0)
        # Opening the file must land on the data, not on the complaints.
        assert load_workbook(buffer).active.title == "Points"

    def test_an_empty_issue_list_adds_no_sheet(self, tmp_path):
        target = tmp_path / "out.xlsx"
        write_workbook(build_table_rows([make_point()]), target, issues=[])
        assert load_workbook(target).sheetnames == ["Points"]


class TestUniquePath:
    """Two exports in the same second must not silently replace each other.
    Before this, a same-minute export destroyed the earlier workbook and both
    runs reported success against the same path."""

    def test_a_free_name_is_used_as_is(self, tmp_path):
        assert unique_path(tmp_path, "points.xlsx") == tmp_path / "points.xlsx"

    def test_a_taken_name_gains_a_counter(self, tmp_path):
        (tmp_path / "points.xlsx").write_text("taken")
        assert unique_path(tmp_path, "points.xlsx") == tmp_path / "points-2.xlsx"

    def test_the_counter_keeps_climbing(self, tmp_path):
        (tmp_path / "points.xlsx").write_text("x")
        (tmp_path / "points-2.xlsx").write_text("x")
        (tmp_path / "points-3.xlsx").write_text("x")
        assert unique_path(tmp_path, "points.xlsx") == tmp_path / "points-4.xlsx"

    def test_the_extension_survives(self, tmp_path):
        (tmp_path / "points.xlsx").write_text("x")
        assert unique_path(tmp_path, "points.xlsx").suffix == ".xlsx"
