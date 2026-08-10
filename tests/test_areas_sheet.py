"""The Areas sheet.

Same three header rows and same 23 columns as Points, so corners read
identically. Each area is a banner carrying its size, then its outer corners,
then a lighter sub-banner and corners for each hole.
"""

import io

import pytest
from openpyxl import load_workbook

from kmz_points.excel import FIRST_DATA_ROW, HEADER_ROW, data_rows, write_workbook
from kmz_points.geometry import measure
from kmz_points.models import Area, Point
from kmz_points.table import (
    COMBINED,
    SEPARATION,
    build_table_rows,
    column_index,
    headers,
)


def corner(lat, lon, name="Plot"):
    return Point(name=name, description="", lon=lon, lat=lat, alt=None, source_file="a.kml")


def box(lat, lon, size, name="Plot"):
    return [
        corner(lat, lon, name),
        corner(lat, lon + size, name),
        corner(lat + size, lon + size, name),
        corner(lat + size, lon, name),
    ]


def make_area(name="Plot 12", lat=10.0, lon=30.0, size=0.01, holes=()):
    return Area(
        name=name,
        description="",
        outer=box(lat, lon, size, name),
        holes=[list(h) for h in holes],
        source_file="a.kml",
    )


def write(areas, points=None):
    """Write a workbook and hand back the reopened Areas sheet (or None)."""
    buffer = io.BytesIO()
    rows = build_table_rows(points or [corner(1.0, 2.0)])
    write_workbook(rows, buffer, areas=[measure(a) for a in areas])
    buffer.seek(0)
    book = load_workbook(buffer)
    return book


def banner_texts(sheet):
    """Every banner on the sheet, in order: rows with column A but no number."""
    number = column_index("#", band=COMBINED)
    texts = []
    for row in sheet.iter_rows(min_row=FIRST_DATA_ROW):
        if row[number].value is None and row[0].value is not None:
            texts.append(str(row[0].value))
    return texts


class TestSheetPresence:
    def test_no_areas_means_no_areas_sheet(self):
        assert write([]).sheetnames == ["Points"]

    def test_areas_get_their_own_sheet_after_points(self):
        assert write([make_area()]).sheetnames == ["Points", "Areas"]

    def test_issues_still_come_last(self):
        buffer = io.BytesIO()
        write_workbook(
            build_table_rows([corner(1.0, 2.0)]),
            buffer,
            issues=["something went wrong"],
            areas=[measure(make_area())],
        )
        buffer.seek(0)
        assert load_workbook(buffer).sheetnames == ["Points", "Areas", "Issues"]

    def test_the_areas_sheet_uses_the_same_columns_as_points(self):
        sheet = write([make_area()])["Areas"]
        assert [c.value for c in sheet[HEADER_ROW]] == headers()


class TestBanners:
    def test_each_area_gets_one_banner(self):
        book = write([make_area("Plot 12"), make_area("Plot 13", lat=20.0)])
        assert len(banner_texts(book["Areas"])) == 2

    def test_the_banner_names_the_area(self):
        assert "Plot 12" in banner_texts(write([make_area("Plot 12")])["Areas"])[0]

    def test_the_banner_carries_all_three_units(self):
        text = banner_texts(write([make_area()])["Areas"])[0]
        assert "m²" in text
        assert "ha" in text
        assert "km²" in text

    def test_the_banner_counts_the_corners(self):
        assert "4 corners" in banner_texts(write([make_area()])["Areas"])[0]

    def test_the_banner_spans_every_column(self):
        sheet = write([make_area()])["Areas"]
        widths = {
            r.max_col - r.min_col + 1
            for r in sheet.merged_cells.ranges
            if r.min_row >= FIRST_DATA_ROW
        }
        assert widths == {len(headers())}


class TestCorners:
    def test_every_corner_becomes_a_row(self):
        assert len(data_rows(write([make_area()])["Areas"])) == 4

    def test_corners_carry_their_coordinates(self):
        sheet = write([make_area(lat=10.0, lon=30.0, size=0.01)])["Areas"]
        longitude = column_index("longitude", band=SEPARATION)
        values = [row[longitude].value for row in data_rows(sheet)]
        assert min(values) == pytest.approx(30.0)
        assert max(values) == pytest.approx(30.01)

    def test_corners_are_numbered_from_one_within_each_area(self):
        sheet = write([make_area("A"), make_area("B", lat=20.0)])["Areas"]
        number = column_index("#", band=COMBINED)
        numbers = [row[number].value for row in data_rows(sheet)]
        assert numbers == [1, 2, 3, 4, 1, 2, 3, 4]

    def test_banner_rows_are_not_counted_as_corners(self):
        # data_rows keys off the numbered column, which banners leave empty.
        sheet = write([make_area(), make_area("Other", lat=20.0)])["Areas"]
        assert len(data_rows(sheet)) == 8


class TestHoles:
    def test_a_hole_gets_its_own_sub_banner(self):
        area = make_area(size=0.02, holes=[box(10.005, 30.005, 0.005)])
        texts = banner_texts(write([area])["Areas"])
        assert len(texts) == 2
        assert "hole 1" in texts[1].lower()

    def test_every_hole_is_announced(self):
        area = make_area(
            size=0.03,
            holes=[box(10.002, 30.002, 0.005), box(10.015, 30.015, 0.005)],
        )
        texts = banner_texts(write([area])["Areas"])
        assert "hole 1" in texts[1].lower()
        assert "hole 2" in texts[2].lower()

    def test_hole_corners_appear_as_rows(self):
        area = make_area(size=0.02, holes=[box(10.005, 30.005, 0.005)])
        assert len(data_rows(write([area])["Areas"])) == 8  # 4 outer + 4 hole

    def test_the_subtracted_size_is_what_the_banner_reports(self):
        outer_only = make_area(size=0.02)
        with_hole = make_area(size=0.02, holes=[box(10.005, 30.005, 0.005)])
        plain = measure(outer_only).measurement.square_metres
        holed = measure(with_hole).measurement.square_metres
        assert holed < plain


class TestUnmeasurableAreas:
    def test_the_banner_says_why_rather_than_showing_a_number(self):
        polar = Area(
            name="Ice",
            description="",
            outer=box(85.0, 30.0, 0.01, "Ice"),
            holes=[],
            source_file="a.kml",
        )
        text = banner_texts(write([polar])["Areas"])[0]
        assert "not measured" in text.lower()
        assert "utm" in text.lower()
        assert "m²" not in text

    def test_its_corners_are_still_listed(self):
        polar = Area(
            name="Ice",
            description="",
            outer=box(85.0, 30.0, 0.01, "Ice"),
            holes=[],
            source_file="a.kml",
        )
        assert len(data_rows(write([polar])["Areas"])) == 4


class TestCornerCount:
    def test_the_repeated_closing_corner_is_not_counted(self):
        # KML writes rings closed. A square has four corners, not five.
        ring = box(10.0, 30.0, 0.01)
        closed = Area(
            name="Square",
            description="",
            outer=ring + [ring[0]],
            holes=[],
            source_file="a.kml",
        )
        assert closed.corner_count == 4
        assert "4 corners" in banner_texts(write([closed])["Areas"])[0]

    def test_an_open_ring_counts_every_corner(self):
        assert make_area().corner_count == 4
