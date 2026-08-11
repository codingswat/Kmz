"""Web service tests.

Flask's test client, so these need no network and no display and run
unchanged on every CI platform.
"""

import io
import re
import tempfile
import threading
import time
from pathlib import Path

import pytest
from openpyxl import load_workbook

from kmz_points.excel import data_rows
from kmz_points.samples import write_samples
from kmz_points.server import LoginAttempts, create_app, safe_upload_name
from kmz_points.table import column_index

PASSWORD = "correct horse"


@pytest.fixture
def client():
    app = create_app(PASSWORD)
    app.config["TESTING"] = True
    return app.test_client()


@pytest.fixture
def signed_in(client):
    client.post("/login", data={"password": PASSWORD})
    return client


@pytest.fixture
def samples(tmp_path):
    return write_samples(tmp_path / "in")


def payload(paths):
    """The multipart body the convert route expects."""
    return {
        "files": [(io.BytesIO(Path(p).read_bytes()), Path(p).name) for p in paths]
    }


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _body_rule(html):
    """The CSS text of the `body { ... }` rule only.

    A plain `"background" in html` substring check would pass for any page
    that declares a background anywhere at all, on any element. Scoping the
    search to the body rule specifically means the test fails if that
    declaration is removed, not just if the stylesheet vanishes entirely.
    """
    match = re.search(r"body\s*\{([^}]*)\}", html)
    assert match, "no body rule found in the page's stylesheet"
    return match.group(1)


def _summary_rule(html):
    """The CSS text of the `.summary { ... }` rule only, scoped the same way
    as `_body_rule` so this cannot be satisfied by a declaration on some
    other, unrelated selector."""
    match = re.search(r"\.summary\s*\{([^}]*)\}", html)
    assert match, "no .summary rule found in the page's stylesheet"
    return match.group(1)


def _declared_background(css_text):
    """The value of a `background:` declaration within a CSS rule's text, or
    None if the rule does not declare one."""
    match = re.search(r"background:\s*([^;]+);", css_text)
    return match.group(1).strip() if match else None


class TestSafeUploadName:
    def test_a_plain_name_is_kept(self):
        assert safe_upload_name("doc.kml", 0) == "doc.kml"

    def test_a_traversal_attempt_loses_its_path(self):
        assert safe_upload_name("../../evil.kml", 0) == "evil.kml"

    def test_a_windows_path_keeps_no_directory_part(self):
        # Asserted as an invariant, not a literal: Path().stem resolves
        # backslashes on Windows but not on POSIX, so the exact string
        # differs by platform while the property that matters does not.
        result = safe_upload_name(r"C:\data\a.kmz", 0)
        assert result.endswith(".kmz")
        assert "/" not in result and "\\" not in result and ":" not in result

    def test_spaces_are_replaced(self):
        assert safe_upload_name("my places.kml", 0) == "my_places.kml"

    def test_a_non_ascii_name_keeps_its_extension(self):
        # secure_filename alone would return "kml", which read_kml_bytes
        # rejects as not a KML file.
        assert safe_upload_name("地図.kml", 3) == "upload_3.kml"

    def test_the_suffix_check_is_case_insensitive(self):
        assert safe_upload_name("DOC.KML", 0) == "DOC.kml"

    def test_a_non_kml_upload_is_refused(self):
        assert safe_upload_name("notes.txt", 0) is None

    def test_a_name_with_no_suffix_is_refused(self):
        assert safe_upload_name("README", 0) is None

    def test_an_absurdly_long_name_is_shortened(self):
        # Filenames over 255 bytes raise OSError 63 on macOS and Linux, which
        # would surface as an unhandled 500 when the upload is saved.
        result = safe_upload_name("x" * 400 + ".kml", 0)
        assert len(result) <= 104
        assert result.endswith(".kml")


class TestPasswordGate:
    def test_an_empty_password_is_refused_at_startup(self):
        with pytest.raises(ValueError):
            create_app("")

    def test_the_front_page_asks_for_the_password(self, client):
        body = client.get("/").get_data(as_text=True)
        assert "password" in body.lower()

    def test_the_wrong_password_is_rejected(self, client):
        response = client.post("/login", data={"password": "wrong"})
        assert response.status_code == 401

    def test_the_right_password_opens_the_upload_page(self, signed_in):
        body = signed_in.get("/").get_data(as_text=True)
        assert ".kml" in body

    def test_the_session_survives_between_requests(self, signed_in):
        assert ".kml" in signed_in.get("/").get_data(as_text=True)
        assert ".kml" in signed_in.get("/").get_data(as_text=True)

    def test_a_non_ascii_password_works(self):
        # hmac.compare_digest refuses str operands that are not both ASCII:
        # "TypeError: comparing strings with non-ASCII characters is not
        # supported". Comparing str would turn every login into a 500 for any
        # owner who picks an accented password, and for any colleague who
        # typos one into the field.
        app = create_app("contraseña")
        app.config["TESTING"] = True
        client = app.test_client()

        assert client.post("/login", data={"password": "wrong"}).status_code == 401
        assert client.post("/login", data={"password": "contraseña"}).status_code == 302

    def test_a_non_ascii_attempt_against_an_ascii_password_is_refused(self, client):
        assert client.post("/login", data={"password": "wrøng"}).status_code == 401

    def test_the_session_cookie_is_locked_down(self, client):
        # Asserted against the real Set-Cookie header from a login through
        # the test client, not app.config: the config keys could be deleted
        # from create_app and this must still catch it, which a check on
        # app.config["SESSION_COOKIE_HTTPONLY"] would not.
        response = client.post("/login", data={"password": PASSWORD})
        cookie = response.headers["Set-Cookie"]
        assert "HttpOnly" in cookie
        assert "SameSite=Lax" in cookie

    def test_the_page_sets_its_own_background_so_dark_mode_is_readable(self, client):
        # Without an explicit body background, the browser supplies its own
        # canvas. In dark mode that canvas is near-black, and the templates'
        # explicit dark ink (#1f2933) renders as dark grey on near-black:
        # barely readable. Checked against a running server in a real
        # browser, not just this test client, which never renders anything.
        login_page = client.get("/").get_data(as_text=True)
        assert "background" in _body_rule(login_page)

        client.post("/login", data={"password": PASSWORD})
        upload_page = client.get("/").get_data(as_text=True)
        assert "background" in _body_rule(upload_page)

    def test_the_summary_block_stays_visually_distinct_from_the_page(self, client):
        # .summary used to stand out against the browser's default canvas on
        # its own #f2f4f7 background. Once body also got an explicit
        # #f2f4f7 background (the dark-mode fix above), the two became
        # byte-identical and the summary block disappeared into the page
        # with nothing to compensate. That block is what a colleague sees
        # when their files failed to convert or produced no points, so
        # losing its separation matters even though the text inside stays
        # legible either way.
        client.post("/login", data={"password": PASSWORD})
        upload_page = client.get("/").get_data(as_text=True)

        body_background = _declared_background(_body_rule(upload_page))
        summary_css = _summary_rule(upload_page)
        summary_background = _declared_background(summary_css)

        distinct_background = (
            summary_background is not None and summary_background != body_background
        )
        # `\bborder\s*:` rather than a plain substring check: `.summary`
        # already declares `border-radius`, which contains the substring
        # "border" but draws nothing by itself without an actual border
        # property to go with it.
        has_border = bool(re.search(r"\bborder\s*:", summary_css))

        assert distinct_background or has_border, (
            ".summary must use a different background than body or declare "
            "a border, or the panel disappears into the page"
        )


class TestConvert:
    def test_converting_needs_a_session(self, client, samples):
        response = client.post("/convert", data=payload(samples))
        assert response.status_code == 302

    def test_the_samples_come_back_as_a_workbook(self, signed_in, samples):
        response = signed_in.post("/convert", data=payload(samples))
        assert response.status_code == 200
        assert response.mimetype == XLSX_MIME
        sheet = load_workbook(io.BytesIO(response.data)).active
        assert len(data_rows(sheet)) == 7  # the 7 sample points

    def test_the_download_is_named_by_the_usual_pattern(self, signed_in, samples):
        response = signed_in.post("/convert", data=payload(samples))
        disposition = response.headers["Content-Disposition"]
        # "attachment" is what makes the browser download the file rather
        # than try to render it inline; "inline; filename=points_....xlsx"
        # would satisfy a check on the filename alone, so that token is
        # asserted explicitly rather than as a side effect of the name check.
        assert disposition.startswith("attachment")
        assert "points_" in disposition and disposition.endswith(".xlsx")

    def test_a_broken_file_does_not_stop_the_good_ones(self, signed_in, samples, tmp_path):
        broken = tmp_path / "broken.kmz"
        broken.write_bytes(b"this is not a zip")
        response = signed_in.post("/convert", data=payload(list(samples) + [broken]))
        assert response.status_code == 200
        assert response.mimetype == XLSX_MIME
        sheet = load_workbook(io.BytesIO(response.data))["Points"]
        assert len(data_rows(sheet)) == 7  # the good files still all came through

    def test_the_failure_is_named_in_the_downloaded_workbook(
        self, signed_in, samples, tmp_path
    ):
        # Not `b"broken.kmz" in response.data`: an xlsx is a zip, so the text
        # is compressed and that check fails even when the sheet is present.
        # Confirmed against a prototype before this test was written.
        broken = tmp_path / "broken.kmz"
        broken.write_bytes(b"this is not a zip")
        response = signed_in.post("/convert", data=payload(list(samples) + [broken]))

        book = load_workbook(io.BytesIO(response.data))
        assert "Issues" in book.sheetnames
        listed = " ".join(
            str(row[0].value) for row in book["Issues"].iter_rows(min_row=2)
        )
        assert "broken.kmz" in listed

    def test_a_batch_with_no_points_returns_a_summary_not_a_download(
        self, signed_in, tmp_path
    ):
        empty = tmp_path / "empty.kml"
        empty.write_text('<kml xmlns="http://www.opengis.net/kml/2.2"><Document/></kml>')
        response = signed_in.post("/convert", data=payload([empty]))
        assert response.status_code == 200
        assert response.mimetype == "text/html"
        assert "No points found" in response.get_data(as_text=True)

    def test_a_non_kml_upload_is_reported_not_crashed(self, signed_in, tmp_path):
        notes = tmp_path / "notes.txt"
        notes.write_text("nothing to see")
        response = signed_in.post("/convert", data=payload([notes]))
        assert response.status_code == 200
        assert "notes.txt" in response.get_data(as_text=True)

    def test_sending_no_files_is_reported(self, signed_in):
        response = signed_in.post("/convert", data={"files": []})
        assert response.status_code == 400
        assert "at least one" in response.get_data(as_text=True).lower()

    def test_an_oversized_upload_is_refused(self, samples):
        app = create_app(PASSWORD, max_upload_bytes=1000)
        app.config["TESTING"] = True
        small = app.test_client()
        small.post("/login", data={"password": PASSWORD})
        response = small.post(
            "/convert", data={"files": [(io.BytesIO(b"x" * 5000), "big.kml")]}
        )
        assert response.status_code == 413

    def test_an_oversized_upload_gets_the_styled_page_not_werkzeugs_bare_one(
        self, samples
    ):
        # Werkzeug's default 413 response names no limit and offers no way
        # back to the form. The registered error handler should replace it
        # with the normal upload page, still answering 413.
        app = create_app(PASSWORD, max_upload_bytes=1000)
        app.config["TESTING"] = True
        small = app.test_client()
        small.post("/login", data={"password": PASSWORD})
        response = small.post(
            "/convert", data={"files": [(io.BytesIO(b"x" * 5000), "big.kml")]}
        )
        body = response.get_data(as_text=True)
        assert response.status_code == 413
        assert "KML / KMZ Point Extractor" in body
        assert "limit" in body

    def test_a_sub_megabyte_cap_is_stated_in_kb_not_as_a_fraction(self):
        # %g rendered a 1000-byte cap as "0.000953674 MB", which reads as
        # gibberish to whoever hit it.
        app = create_app(PASSWORD, max_upload_bytes=1000)
        app.config["TESTING"] = True
        client = app.test_client()
        client.post("/login", data={"password": PASSWORD})
        body = client.get("/").get_data(as_text=True)
        assert "KB" in body
        assert "0.00" not in body

    def test_the_upload_page_states_the_size_limit_up_front(self, signed_in):
        body = signed_in.get("/").get_data(as_text=True)
        assert "50 MB" in body

    def test_nothing_is_left_behind_on_disk(self, signed_in, samples, monkeypatch, tmp_path):
        # tempfile.tempdir is pinned to a directory only this test uses.
        # Scanning the shared system temp directory instead would fail
        # whenever any unrelated process happened to create a file mid-test.
        scratch = tmp_path / "scratch"
        scratch.mkdir()
        monkeypatch.setattr(tempfile, "tempdir", str(scratch))

        signed_in.post("/convert", data=payload(samples))

        assert list(scratch.iterdir()) == []

    def test_a_traversal_filename_cannot_escape_the_workspace(
        self, signed_in, samples, monkeypatch, tmp_path
    ):
        # The same check as safe_upload_name's unit test, but driven through
        # the real route, which is where it actually matters.
        scratch = tmp_path / "scratch"
        scratch.mkdir()
        monkeypatch.setattr(tempfile, "tempdir", str(scratch))

        body = Path(samples[0]).read_bytes()
        response = signed_in.post(
            "/convert", data={"files": [(io.BytesIO(body), "../../evil.kml")]}
        )

        assert response.status_code == 200
        assert list(scratch.iterdir()) == []
        assert not (tmp_path / "evil.kml").exists()
        assert not (tmp_path.parent / "evil.kml").exists()

    def test_two_uploads_with_the_same_name_do_not_collide(self, signed_in):
        # "doc.kml" twice: both sanitise to the identical stem, so without a
        # per-upload subdirectory the second save can overwrite the first on
        # disk before both are accounted for.
        first = (
            b'<?xml version="1.0"?>'
            b'<kml xmlns="http://www.opengis.net/kml/2.2"><Document>'
            b"<Placemark><name>Alpha</name>"
            b"<Point><coordinates>1,2</coordinates></Point></Placemark>"
            b"</Document></kml>"
        )
        second = (
            b'<?xml version="1.0"?>'
            b'<kml xmlns="http://www.opengis.net/kml/2.2"><Document>'
            b"<Placemark><name>Bravo</name>"
            b"<Point><coordinates>3,4</coordinates></Point></Placemark>"
            b"</Document></kml>"
        )

        response = signed_in.post(
            "/convert",
            data={
                "files": [
                    (io.BytesIO(first), "doc.kml"),
                    (io.BytesIO(second), "doc.kml"),
                ]
            },
        )

        assert response.status_code == 200
        assert response.mimetype == XLSX_MIME
        sheet = load_workbook(io.BytesIO(response.data))["Points"]
        # A row count alone cannot tell a healthy batch from one where the
        # first upload was clobbered and the second double-counted: both
        # shapes have 2 data rows. Check the actual point names instead.
        # data_rows (not a raw min_row scan) is required here too: both
        # uploads sanitise to the same stem, so _write_body sees the same
        # source_file value twice in a row and groups them under a single
        # shared grey banner naming "doc.kml" rather than one per point --
        # and that banner would otherwise land in the same column as Name.
        name_column = column_index("Name")
        names = {row[name_column].value for row in data_rows(sheet)}
        assert names == {"Alpha", "Bravo"}


AREAS_ONLY_KML = b"""<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2"><Document>
  <Placemark><name>Plot 12</name><Polygon><outerBoundaryIs><LinearRing>
    <coordinates>38.20,34.60 38.21,34.60 38.21,34.61 38.20,34.61 38.20,34.60</coordinates>
  </LinearRing></outerBoundaryIs></Polygon></Placemark>
</Document></kml>"""


class TestAreasOnlyBatch:
    """A file holding shapes but no points still has something to export."""

    def test_a_batch_of_only_areas_returns_a_workbook(self, signed_in):
        response = signed_in.post(
            "/convert",
            data={"files": [(io.BytesIO(AREAS_ONLY_KML), "plots.kml")]},
        )
        assert response.status_code == 200
        assert response.mimetype == XLSX_MIME

    def test_that_workbook_carries_the_areas_sheet(self, signed_in):
        response = signed_in.post(
            "/convert",
            data={"files": [(io.BytesIO(AREAS_ONLY_KML), "plots.kml")]},
        )
        book = load_workbook(io.BytesIO(response.data))
        assert "Areas" in book.sheetnames

    def test_a_batch_with_neither_points_nor_areas_still_reports(self, signed_in):
        empty = b'<kml xmlns="http://www.opengis.net/kml/2.2"><Document/></kml>'
        response = signed_in.post(
            "/convert", data={"files": [(io.BytesIO(empty), "empty.kml")]}
        )
        assert response.status_code == 200
        assert response.mimetype == "text/html"
        assert "No points found" in response.get_data(as_text=True)


class TestConvertFeedback:
    """A successful conversion is a download, so the page never navigates and
    nothing tells it the work finished. Without a signal the Convert button
    looks idle while a large batch is still being read, and people submit it
    twice."""

    def test_the_page_offers_a_working_indicator(self, signed_in):
        body = signed_in.get("/").get_data(as_text=True)
        assert 'id="convert-button"' in body
        assert "Converting" in body

    def test_the_download_echoes_the_token_back_as_a_cookie(self, signed_in, samples):
        data = payload(samples)
        data["download_token"] = "abc123"
        response = signed_in.post("/convert", data=data)
        assert response.status_code == 200
        assert "download_token=abc123" in response.headers.get("Set-Cookie", "")

    def test_the_cookie_is_readable_by_the_page(self, signed_in, samples):
        # HttpOnly would hide it from the script that has to see it.
        data = payload(samples)
        data["download_token"] = "abc123"
        response = signed_in.post("/convert", data=data)
        assert "HttpOnly" not in response.headers.get("Set-Cookie", "")

    def test_no_token_means_no_cookie(self, signed_in, samples):
        response = signed_in.post("/convert", data=payload(samples))
        assert "download_token" not in response.headers.get("Set-Cookie", "")

    def test_the_form_still_submits_without_javascript(self, signed_in, samples):
        # The script only decorates; the plain form post must still convert.
        response = signed_in.post("/convert", data=payload(samples))
        assert response.status_code == 200
        assert response.mimetype == XLSX_MIME


class TestAreasOnThePage:
    def test_the_page_says_what_happens_to_areas(self, signed_in):
        body = signed_in.get("/").get_data(as_text=True)
        assert "Areas" in body or "areas" in body
        assert "km²" in body

    def test_the_summary_reports_the_area_count(self, signed_in):
        empty = b'<kml xmlns="http://www.opengis.net/kml/2.2"><Document/></kml>'
        response = signed_in.post(
            "/convert", data={"files": [(io.BytesIO(empty), "empty.kml")]}
        )
        assert "area(s) extracted" in response.get_data(as_text=True)


class TestDropZone:
    """Drag-and-drop is decoration over the file input, not a replacement:
    the zone is hidden until the script reveals it, so a browser without
    JavaScript never shows an invitation it cannot honour."""

    def test_the_zone_ships_hidden(self, signed_in):
        body = signed_in.get("/").get_data(as_text=True)
        assert 'id="drop-zone" hidden' in body

    def test_the_file_input_is_still_present_and_required(self, signed_in):
        body = signed_in.get("/").get_data(as_text=True)
        assert 'type="file"' in body
        assert "required" in body

    def test_the_zone_says_what_it_accepts(self, signed_in):
        body = signed_in.get("/").get_data(as_text=True)
        assert ".kml" in body and ".kmz" in body

    def test_the_page_still_posts_without_the_script(self, signed_in, samples):
        # Whatever the script does, the plain form post must still convert.
        response = signed_in.post("/convert", data=payload(samples))
        assert response.status_code == 200
        assert response.mimetype == XLSX_MIME


class TestLoginThrottle:
    """Guessing the one shared password must not be free.

    The fixtures above build a fresh app per test, so the counter starts empty
    every time and these cannot make the suite order-dependent -- which also
    means the existing "wrong password gives 401" tests keep getting 401, as
    they are nowhere near the limit.
    """

    def test_the_first_attempts_are_refused_the_ordinary_way(self, client):
        for _ in range(5):
            assert client.post("/login", data={"password": "wrong"}).status_code == 401

    def test_one_attempt_past_the_limit_is_throttled(self, client):
        for _ in range(5):
            client.post("/login", data={"password": "wrong"})
        response = client.post("/login", data={"password": "wrong"})
        assert response.status_code == 429

    def test_the_throttle_page_says_how_long_to_wait(self, client):
        for _ in range(6):
            response = client.post("/login", data={"password": "wrong"})
        body = response.get_data(as_text=True)
        assert "Too many" in body
        assert "seconds" in body

    def test_the_right_password_is_refused_once_throttled(self, client):
        # The point of a throttle is that it does not care whether this one
        # happens to be right; otherwise guessing costs nothing again.
        for _ in range(5):
            client.post("/login", data={"password": "wrong"})
        assert client.post("/login", data={"password": PASSWORD}).status_code == 429

    def test_a_success_before_the_limit_clears_the_count(self, client):
        for _ in range(4):
            client.post("/login", data={"password": "wrong"})
        assert client.post("/login", data={"password": PASSWORD}).status_code == 302

        # Four more would have tripped the old count; they must not now.
        for _ in range(4):
            assert client.post("/login", data={"password": "wrong"}).status_code == 401

    def test_the_window_expires(self):
        app = create_app(PASSWORD, attempt_window=0.05)
        app.config["TESTING"] = True
        client = app.test_client()

        for _ in range(5):
            client.post("/login", data={"password": "wrong"})
        assert client.post("/login", data={"password": "wrong"}).status_code == 429

        time.sleep(0.06)
        assert client.post("/login", data={"password": "wrong"}).status_code == 401

    def test_the_limit_is_configurable(self):
        app = create_app(PASSWORD, max_attempts=2)
        app.config["TESTING"] = True
        client = app.test_client()

        assert client.post("/login", data={"password": "wrong"}).status_code == 401
        assert client.post("/login", data={"password": "wrong"}).status_code == 401
        assert client.post("/login", data={"password": "wrong"}).status_code == 429


class TestAttemptTableBounds:
    """The counter must not become a slow leak keyed by anything that connects."""

    def test_it_holds_one_entry_per_recently_failing_address(self):
        attempts = LoginAttempts(limit=5, window=60.0)
        for index in range(50):
            attempts.record_failure(f"10.0.0.{index}")
        assert attempts.tracked() == 50

    def test_addresses_whose_failures_aged_out_are_dropped(self):
        attempts = LoginAttempts(limit=5, window=0.05)
        for index in range(50):
            attempts.record_failure(f"10.0.0.{index}")
        assert attempts.tracked() == 50

        time.sleep(0.06)
        # Eviction runs on write, so one more failure is what prunes the rest.
        # The table then holds exactly that one address -- the bound is "one
        # entry per address that failed inside the window", and after the
        # sleep only this address has.
        attempts.record_failure("10.0.0.99")
        assert attempts.tracked() == 1

    def test_a_cleared_address_is_forgotten(self):
        attempts = LoginAttempts()
        attempts.record_failure("10.0.0.1")
        assert attempts.tracked() == 1
        attempts.clear("10.0.0.1")
        assert attempts.tracked() == 0

    def test_counting_is_safe_from_several_threads(self):
        # serve.py runs with eight worker threads, so the counter is shared.
        attempts = LoginAttempts(limit=1000, window=60.0)

        def hammer():
            for _ in range(200):
                attempts.record_failure("10.0.0.1")

        workers = [threading.Thread(target=hammer) for _ in range(8)]
        for worker in workers:
            worker.start()
        for worker in workers:
            worker.join()

        assert attempts.blocked("10.0.0.1") is True
        assert attempts.tracked() == 1
