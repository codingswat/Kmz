"""Pipeline tests, including a full end-to-end run over generated samples."""

import io
from pathlib import Path

import openpyxl
import pytest

from kmz_points.excel import HEADER_ROW, data_rows
from kmz_points.pipeline import export_to_excel, export_to_stream, load_file, run
from kmz_points.samples import write_samples
from kmz_points.table import COMBINED, column_index, headers

# What write_samples() is defined to produce.
SAMPLE_POINT_TOTAL = 7
# One LineString route. The sample Polygon used to be skipped too; it is
# extracted as an area now.
SAMPLE_SKIPPED_TOTAL = 1


@pytest.fixture
def samples(tmp_path):
    return write_samples(tmp_path / "samples")


class TestWriteSamples:
    def test_creates_the_three_sample_inputs(self, samples):
        assert sorted(p.name for p in samples) == [
            "nested.kml",
            "sample.kmz",
            "simple.kml",
        ]

    def test_samples_exist_on_disk(self, samples):
        assert all(p.is_file() for p in samples)


class TestLoadFile:
    def test_loads_a_plain_kml(self, samples):
        simple = next(p for p in samples if p.name == "simple.kml")
        result = load_file(simple)
        assert result.point_count == 2
        assert result.error is None

    def test_loads_a_kmz(self, samples):
        kmz = next(p for p in samples if p.name == "sample.kmz")
        result = load_file(kmz)
        assert result.point_count == 2
        assert result.error is None

    def test_counts_non_point_features_without_extracting_them(self, samples):
        nested = next(p for p in samples if p.name == "nested.kml")
        result = load_file(nested)
        assert result.point_count == 3
        assert result.skipped == 1  # the route; the polygon is now an area

    def test_points_are_attributed_to_their_source_file(self, samples):
        simple = next(p for p in samples if p.name == "simple.kml")
        assert {p.source_file for p in load_file(simple).points} == {"simple.kml"}

    def test_missing_file_is_reported_not_raised(self, tmp_path):
        result = load_file(tmp_path / "absent.kml")
        assert result.error is not None
        assert result.points == []

    def test_kmz_without_kml_is_reported_not_raised(self, tmp_path):
        import zipfile

        broken = tmp_path / "empty.kmz"
        with zipfile.ZipFile(broken, "w") as archive:
            archive.writestr("readme.txt", "nothing")
        assert load_file(broken).error is not None

    def test_file_with_zero_points_is_not_an_error(self, tmp_path):
        empty = tmp_path / "empty.kml"
        empty.write_text('<kml xmlns="http://www.opengis.net/kml/2.2"><Document/></kml>')
        result = load_file(empty)
        assert result.error is None
        assert result.point_count == 0


class TestExport:
    def test_writes_a_workbook_named_by_pattern(self, samples, tmp_path):
        loaded = [load_file(p) for p in samples]
        summary = export_to_excel(loaded, tmp_path)
        assert summary.output_path is not None
        # Path().name rather than splitting on "/" -- Windows separates with a
        # backslash, so the split returned the whole path and the assertion
        # compared "points_" against a drive letter.
        name = Path(summary.output_path).name
        assert name.startswith("points_") and name.endswith(".xlsx")

    def test_summary_counts_every_file_and_point(self, samples, tmp_path):
        summary = export_to_excel([load_file(p) for p in samples], tmp_path)
        assert summary.files_read == 3
        assert summary.points_extracted == SAMPLE_POINT_TOTAL
        assert summary.features_skipped == SAMPLE_SKIPPED_TOTAL
        assert summary.files_failed == 0

    def test_failed_files_are_counted_and_do_not_stop_the_export(
        self, samples, tmp_path
    ):
        loaded = [load_file(p) for p in samples] + [load_file(tmp_path / "absent.kml")]
        summary = export_to_excel(loaded, tmp_path)
        assert summary.files_failed == 1
        assert summary.points_extracted == SAMPLE_POINT_TOTAL
        assert summary.output_path is not None

    def test_export_with_no_points_writes_nothing_and_says_so(self, tmp_path):
        summary = export_to_excel([], tmp_path)
        assert summary.output_path is None
        assert summary.warnings


class TestEndToEnd:
    def test_run_produces_a_readable_workbook(self, samples, tmp_path):
        out = tmp_path / "out"
        summary = run(samples, out)
        book = openpyxl.load_workbook(summary.output_path)
        assert [c.value for c in book.active[HEADER_ROW]] == headers()

    def test_every_extracted_point_becomes_a_row(self, samples, tmp_path):
        summary = run(samples, tmp_path / "out")
        book = openpyxl.load_workbook(summary.output_path)
        assert len(data_rows(book.active)) == SAMPLE_POINT_TOTAL

    def test_numbering_runs_unbroken_across_the_whole_batch(self, samples, tmp_path):
        summary = run(samples, tmp_path / "out")
        sheet = openpyxl.load_workbook(summary.output_path).active
        number = column_index("#", band=COMBINED)
        numbers = [row[number].value for row in data_rows(sheet)]
        assert numbers == list(range(1, SAMPLE_POINT_TOTAL + 1))

    def test_no_mandatory_cell_is_left_empty(self, samples, tmp_path):
        summary = run(samples, tmp_path / "out")
        sheet = openpyxl.load_workbook(summary.output_path).active
        # Elevation and Description are legitimately optional; everything
        # else must be populated for every row.
        optional = {"elevation", "Description"}
        required = [i for i, h in enumerate(headers()) if h not in optional]
        for row in data_rows(sheet):
            for index in required:
                value = row[index].value
                assert value not in (None, ""), (
                    f"row {row[0].row} column {headers()[index]!r} is empty"
                )

    def test_all_three_source_files_are_represented(self, samples, tmp_path):
        summary = run(samples, tmp_path / "out")
        sheet = openpyxl.load_workbook(summary.output_path).active
        source = column_index("Source File")
        sources = {row[source].value for row in data_rows(sheet)}
        assert sources == {"simple.kml", "nested.kml", "sample.kmz"}

    def test_output_lands_in_the_requested_directory(self, samples, tmp_path):
        out = tmp_path / "chosen"
        summary = run(samples, out)
        assert summary.output_path.startswith(str(out))

    def test_each_source_file_gets_its_own_banner_in_order(self, samples, tmp_path):
        # An end-to-end guard for the banner grouping: it relies on points
        # arriving in file order, which is a property of the real pipeline
        # that a synthetic, hand-ordered test can't accidentally violate.
        summary = run(samples, tmp_path / "out")
        sheet = openpyxl.load_workbook(summary.output_path).active
        source = column_index("Source File")

        seen = []
        previous = None
        for row in data_rows(sheet):
            current = row[source].value
            if current != previous:
                banner_row = row[0].row - 1
                assert sheet.cell(row=banner_row, column=1).value == current
                seen.append(current)
                previous = current
        assert seen == ["simple.kml", "nested.kml", "sample.kmz"]


class TestExportToStream:
    """The web service path: same batch, no file on disk."""

    def test_the_stream_holds_a_readable_workbook(self, samples, tmp_path):
        loaded = [load_file(p) for p in samples]
        buffer = io.BytesIO()
        export_to_stream(loaded, buffer)
        buffer.seek(0)
        sheet = openpyxl.load_workbook(buffer).active
        assert len(data_rows(sheet)) == SAMPLE_POINT_TOTAL

    def test_the_summary_matches_the_file_based_export(self, samples, tmp_path):
        # Guards the _collect split: the two paths must not drift.
        loaded = [load_file(p) for p in samples]
        to_file = export_to_excel([load_file(p) for p in samples], tmp_path)
        to_stream = export_to_stream(loaded, io.BytesIO())

        assert to_stream.files_read == to_file.files_read
        assert to_stream.files_failed == to_file.files_failed
        assert to_stream.points_extracted == to_file.points_extracted
        assert to_stream.features_skipped == to_file.features_skipped
        assert to_stream.warnings == to_file.warnings

    def test_no_points_writes_nothing_to_the_stream(self, tmp_path):
        empty = tmp_path / "empty.kml"
        empty.write_text('<kml xmlns="http://www.opengis.net/kml/2.2"><Document/></kml>')
        buffer = io.BytesIO()
        summary = export_to_stream([load_file(empty)], buffer)
        assert summary.points_extracted == 0
        assert buffer.getvalue() == b""

    def test_output_path_is_never_set_for_a_stream(self, samples):
        loaded = [load_file(p) for p in samples]
        assert export_to_stream(loaded, io.BytesIO()).output_path is None

    def test_a_failed_file_is_named_inside_the_workbook(self, samples, tmp_path):
        # The browser gets the file and nothing else, so this is the only
        # place a partial failure can be reported.
        broken = tmp_path / "broken.kmz"
        broken.write_bytes(b"this is not a zip")
        loaded = [load_file(p) for p in list(samples) + [broken]]

        buffer = io.BytesIO()
        summary = export_to_stream(loaded, buffer)
        buffer.seek(0)
        book = openpyxl.load_workbook(buffer)

        assert summary.points_extracted == SAMPLE_POINT_TOTAL
        assert "Issues" in book.sheetnames
        listed = " ".join(
            str(row[0].value) for row in book["Issues"].iter_rows(min_row=2)
        )
        assert "broken.kmz" in listed

    def test_a_clean_batch_has_no_issues_sheet(self, samples):
        loaded = [load_file(p) for p in samples]
        buffer = io.BytesIO()
        export_to_stream(loaded, buffer)
        buffer.seek(0)
        # Not an exact sheet list: the samples contain a polygon, so an Areas
        # sheet is expected. What must be absent is Issues.
        assert "Issues" not in openpyxl.load_workbook(buffer).sheetnames

    def test_the_file_export_gains_no_issues_sheet(self, samples, tmp_path):
        # The desktop app's output must not change shape.
        broken = tmp_path / "broken.kmz"
        broken.write_bytes(b"this is not a zip")
        loaded = [load_file(p) for p in list(samples) + [broken]]

        summary = export_to_excel(loaded, tmp_path)
        sheets = openpyxl.load_workbook(summary.output_path).sheetnames
        assert "Issues" not in sheets


class TestRepeatedExportsDoNotOverwrite:
    """A second export in the same second used to replace the first one's
    workbook on disk while both runs reported success against the same path,
    so the earlier batch was lost with nothing to indicate it."""

    def test_two_exports_in_the_same_second_produce_two_files(
        self, samples, tmp_path
    ):
        from datetime import datetime

        out = tmp_path / "out"
        out.mkdir()
        when = datetime(2026, 8, 10, 14, 30, 15)

        first = export_to_excel([load_file(p) for p in samples], out, when)
        second = export_to_excel([load_file(samples[0])], out, when)

        assert first.output_path != second.output_path
        assert len(list(out.glob("points_*.xlsx"))) == 2

    def test_the_earlier_workbook_still_holds_its_own_points(
        self, samples, tmp_path
    ):
        from datetime import datetime

        out = tmp_path / "out"
        out.mkdir()
        when = datetime(2026, 8, 10, 14, 30, 15)

        first = export_to_excel([load_file(p) for p in samples], out, when)
        export_to_excel([load_file(samples[0])], out, when)

        book = openpyxl.load_workbook(first.output_path)
        assert len(data_rows(book.active)) == SAMPLE_POINT_TOTAL

    def test_the_second_file_is_named_by_a_counter(self, samples, tmp_path):
        from datetime import datetime

        out = tmp_path / "out"
        out.mkdir()
        when = datetime(2026, 8, 10, 14, 30, 15)

        export_to_excel([load_file(p) for p in samples], out, when)
        second = export_to_excel([load_file(samples[0])], out, when)

        assert Path(second.output_path).name.endswith("-2.xlsx")
