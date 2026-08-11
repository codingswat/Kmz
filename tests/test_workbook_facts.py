"""What workbook_facts() reports about a workbook it is handed.

The agreement between the Python and browser writers is only as good as this
reduction: a fact it quietly drops is a divergence neither suite can see, so
each one it claims to report is checked here against a workbook small enough
to read in full.
"""

import io

import openpyxl
import pytest

from kmz_points.workbook_facts import workbook_facts


def sample_book():
    """A workbook carrying one of everything the facts describe."""
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Sample"

    sheet["A1"] = "Report"
    # Merged out of order, so the reported order is the sort and not the
    # order openpyxl happens to hand them back in.
    sheet.merge_cells("A5:C5")
    sheet.merge_cells("A1:C1")

    sheet["A2"] = "Alpha"
    sheet["B2"] = 34.5678904999
    sheet["B2"].number_format = "0.000000"
    sheet["C2"] = 12

    sheet.freeze_panes = "A3"
    for letter, width in (("A", 12), ("B", 25.5), ("C", 8.129)):
        sheet.column_dimensions[letter].width = width

    plain = book.create_sheet("Plain")
    plain["A1"] = "nothing frozen, nothing merged"
    plain.column_dimensions["A"].width = 30
    return book


@pytest.fixture
def facts(tmp_path):
    """Facts read back from a saved file, which is how they are really used."""
    path = tmp_path / "sample.xlsx"
    sample_book().save(path)
    return workbook_facts(path)


@pytest.fixture
def sheet(facts):
    return facts["sheets"][0]


def cell_at(sheet, row, col):
    return next(c for c in sheet["cells"] if (c["row"], c["col"]) == (row, col))


class TestSheets:
    def test_every_sheet_is_reported_in_order(self, facts):
        assert [s["name"] for s in facts["sheets"]] == ["Sample", "Plain"]


class TestFreezePanes:
    def test_a_frozen_pane_becomes_the_last_frozen_row(self, sheet):
        # openpyxl says "A3", meaning rows 1 and 2 are frozen.
        assert sheet["freezeRow"] == 2

    def test_a_sheet_that_freezes_nothing_reports_zero(self, facts):
        assert facts["sheets"][1]["freezeRow"] == 0


class TestMerges:
    def test_merged_ranges_are_listed_in_sorted_order(self, sheet):
        assert sheet["merges"] == ["A1:C1", "A5:C5"]

    def test_a_sheet_without_merges_lists_none(self, facts):
        assert facts["sheets"][1]["merges"] == []


class TestColumnWidths:
    def test_one_width_per_column_the_sheet_uses(self, sheet):
        assert len(sheet["columnWidths"]) == 3

    def test_widths_are_rounded_to_two_places(self, sheet):
        assert sheet["columnWidths"] == [12, 25.5, 8.13]


class TestCells:
    def test_a_string_cell(self, sheet):
        assert cell_at(sheet, 2, 1) == {
            "row": 2,
            "col": 1,
            "value": "Alpha",
            "type": "s",
            "numberFormat": "General",
        }

    def test_a_number_cell_carries_its_format(self, sheet):
        assert cell_at(sheet, 2, 2) == {
            "row": 2,
            "col": 2,
            "value": 34.56789,
            "type": "n",
            "numberFormat": "0.000000",
        }

    def test_numbers_are_rounded_to_six_places(self, sheet):
        # Written as 34.5678904999: the tail is float noise of the kind two
        # languages computing the same number disagree about, not a value.
        assert cell_at(sheet, 2, 2)["value"] == 34.56789

    def test_an_integer_is_still_a_number(self, sheet):
        assert (cell_at(sheet, 2, 3)["value"], cell_at(sheet, 2, 3)["type"]) == (12, "n")

    def test_empty_cells_are_left_out(self, sheet):
        # Four cells hold something; B1 and C1 were swallowed by the merge,
        # and row 5 is a merge over nothing at all.
        assert len(sheet["cells"]) == 4

    def test_cells_are_ordered_by_row_then_column(self, sheet):
        positions = [(c["row"], c["col"]) for c in sheet["cells"]]
        assert positions == sorted(positions)


class TestSources:
    """A path, a stream and a live Workbook must describe the same workbook."""

    def test_a_stream_reads_the_same_as_a_path(self, facts):
        buffer = io.BytesIO()
        sample_book().save(buffer)
        buffer.seek(0)
        assert workbook_facts(buffer) == facts

    def test_a_workbook_object_needs_no_saving(self, facts):
        assert workbook_facts(sample_book()) == facts
